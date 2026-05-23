"""
StudyPilot — Export Engine
Handles Excel, CSV exports of timetable grids.
"""

import io
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Color palette for Excel cell fills
FILL_COLORS = {
    "📘": "C3B1E1",   # college — lavender
    "🏆": "FFD700",   # competitive — gold
    "🔁": "90EE90",   # revision — light green
    "☕": "FFE4B5",   # break — moccasin
    "😴": "B0C4DE",   # sleep — steel blue
    "🎮": "98FB98",   # free — pale green
    "🏃": "87CEFA",   # exercise — sky blue
    "🍽️": "FFA07A",   # meal — salmon
    "🚫": "D3D3D3",   # blocked — light grey
    "🎓": "DDA0DD",   # college class — plum
}

DEFAULT_FILL = "FFFFFF"

WEEK = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def _get_fill(label: str) -> str:
    for emoji, color in FILL_COLORS.items():
        if label.startswith(emoji):
            return color
    return DEFAULT_FILL


def export_to_excel(grids: List[Dict], name: str = "StudyPilot_Timetable") -> bytes:
    """Export multi-week grids to Excel workbook bytes."""
    wb = Workbook()
    wb.remove(wb.active)

    hours = [f"{h:02d}:00-{(h+1)%24:02d}:00" for h in range(24)]
    header_fill = PatternFill("solid", fgColor="6C63FF")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for w, grid in enumerate(grids, start=1):
        ws = wb.create_sheet(title=f"Week {w}")
        ws.column_dimensions["A"].width = 14

        # Header row
        ws.cell(row=1, column=1, value="Day / Time")
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).font = header_font

        for col, hr in enumerate(hours, start=2):
            cell = ws.cell(row=1, column=col, value=hr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 22

        # Data rows
        for row_idx, day in enumerate(WEEK, start=2):
            ws.cell(row=row_idx, column=1, value=day).font = Font(bold=True)
            day_slots = grid.get(day, {})
            for col, hr in enumerate(hours, start=2):
                label = day_slots.get(hr, "")
                cell = ws.cell(row=row_idx, column=col, value=label)
                fill_color = _get_fill(label)
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                cell.border = border
                ws.row_dimensions[row_idx].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_to_csv(grid: Dict) -> bytes:
    """Export single-week grid to CSV bytes."""
    hours = [f"{h:02d}:00-{(h+1)%24:02d}:00" for h in range(24)]
    rows = []
    for day in WEEK:
        row = {"Day": day}
        for hr in hours:
            row[hr] = grid.get(day, {}).get(hr, "")
        rows.append(row)
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    df.to_csv(buf, index=False)
    return buf.getvalue()
